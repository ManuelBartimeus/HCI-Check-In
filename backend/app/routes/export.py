"""
Export routes — attendance CSV/Excel + leaderboard export.
"""
import io
import csv
from datetime import datetime
from flask import Blueprint, request, jsonify, make_response
from flask_jwt_extended import jwt_required
from ..models.attendance import Attendance
from ..models.member import Member
from ..services.leaderboard_service import get_leaderboard

export_bp = Blueprint('export', __name__)


def _attendance_query(args):
    q = Attendance.query.order_by(Attendance.timestamp.desc())
    if args.get('member_id'):
        q = q.filter(Attendance.member_id == int(args['member_id']))
    if args.get('meeting_id'):
        q = q.filter(Attendance.meeting_id == int(args['meeting_id']))
    if args.get('date_from'):
        try:
            q = q.filter(Attendance.timestamp >= datetime.strptime(args['date_from'], '%Y-%m-%d'))
        except ValueError:
            pass
    if args.get('date_to'):
        try:
            q = q.filter(Attendance.timestamp <= datetime.strptime(args['date_to'], '%Y-%m-%d').replace(hour=23, minute=59))
        except ValueError:
            pass
    return q


@export_bp.route('/api/export/attendance/csv', methods=['GET'])
@jwt_required()
def export_attendance_csv():
    """Export filtered attendance as CSV."""
    rows = _attendance_query(request.args).all()

    output = io.StringIO()
    writer = csv.writer(output)
    writer.writerow(['ID', 'Member', 'Meeting', 'Timestamp', 'Points Awarded', 'Status', 'Source'])
    for a in rows:
        writer.writerow([
            a.id,
            a.member.name if a.member else '',
            a.meeting.name if a.meeting else '',
            a.timestamp.strftime('%Y-%m-%d %H:%M:%S'),
            a.points_awarded,
            a.status,
            a.source,
        ])

    response = make_response(output.getvalue())
    response.headers['Content-Type'] = 'text/csv'
    response.headers['Content-Disposition'] = (
        f'attachment; filename=attendance_{datetime.utcnow().strftime("%Y%m%d_%H%M%S")}.csv'
    )
    return response


@export_bp.route('/api/export/attendance/excel', methods=['GET'])
@jwt_required()
def export_attendance_excel():
    """Export filtered attendance as Excel .xlsx."""
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment
    except ImportError:
        return jsonify({'error': 'openpyxl not installed'}), 500

    rows = _attendance_query(request.args).all()

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Attendance'

    headers = ['ID', 'Member', 'Meeting', 'Date', 'Time', 'Points Awarded', 'Status', 'Source']
    header_font = Font(bold=True, color='FFFFFF')
    header_fill = PatternFill('solid', fgColor='141414')

    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center')

    for row_idx, a in enumerate(rows, 2):
        ws.cell(row=row_idx, column=1, value=a.id)
        ws.cell(row=row_idx, column=2, value=a.member.name if a.member else '')
        ws.cell(row=row_idx, column=3, value=a.meeting.name if a.meeting else '')
        ws.cell(row=row_idx, column=4, value=a.timestamp.strftime('%Y-%m-%d'))
        ws.cell(row=row_idx, column=5, value=a.timestamp.strftime('%H:%M:%S'))
        ws.cell(row=row_idx, column=6, value=a.points_awarded)
        ws.cell(row=row_idx, column=7, value=a.status)
        ws.cell(row=row_idx, column=8, value=a.source)

    # Auto-size columns
    for col in ws.columns:
        max_len = max((len(str(c.value or '')) for c in col), default=10)
        ws.column_dimensions[col[0].column_letter].width = min(max_len + 4, 40)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    response = make_response(buf.read())
    response.headers['Content-Type'] = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    response.headers['Content-Disposition'] = (
        f'attachment; filename=attendance_{datetime.utcnow().strftime("%Y%m%d_%H%M%S")}.xlsx'
    )
    return response


@export_bp.route('/api/export/leaderboard/csv', methods=['GET'])
@jwt_required()
def export_leaderboard_csv():
    """Export leaderboard as CSV."""
    data = get_leaderboard()

    output = io.StringIO()
    writer = csv.writer(output)
    writer.writerow(['Rank', 'Name', 'Points'])
    for entry in data['leaderboard']:
        writer.writerow([entry['rank'], entry['name'], entry['points']])

    response = make_response(output.getvalue())
    response.headers['Content-Type'] = 'text/csv'
    response.headers['Content-Disposition'] = (
        f'attachment; filename=leaderboard_{datetime.utcnow().strftime("%Y%m%d_%H%M%S")}.csv'
    )
    return response
